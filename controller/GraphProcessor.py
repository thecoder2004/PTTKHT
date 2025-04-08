import os
import re
import json
from controller.NodeGenerator import TimeoutNode
from controller.NodeGenerator import ArtificialNode
from controller.NodeGenerator import TimeWindowNode
from controller.NodeGenerator import RestrictionNode
from controller.RestrictionController import RestrictionController
from model.Node import Node
from model.hallway_simulator_module.HallwaySimulator import BulkHallwaySimulator
from collections import deque
from scipy.sparse import lil_matrix
import numpy as np
import pdb
import config

""" Mô tả yêu cầu của code:
https://docs.google.com/document/d/13S_Ycg-aB4GjEm8xe6tAoUHzhS-Z1iFnM4jX_bWFddo/edit?usp=sharing """

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    RED = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

class GraphProcessor:
    def __init__(self):
        self._adj = []  # Adjacency matrix
        self._M = 0
        self._H = 0
        self._draw = 0
        self._d = 0
        self._alpha = 1
        self._beta = 1
        self._gamma = 1
        self._ID = []
        self._earliness = 0
        self._target_nodes = []
        self._tardiness = 0
        self._space_edges = []
        self._ts_edges = []
        self._ts_nodes = []
        self._tsedges = []
        self._started_nodes = []
        self._print_out = True
        self._time_window_controller = None
        self._restriction_controller = None
        self._start_ban = -1
        self._end_ban = -1
        self._seed = 0
        self._num_max_agvs = 0
        self._graph = None
        if(config.level_of_simulation == 1):
            #random in the list
            #import os
            # Lấy thư mục hiện hành
            #current_directory = os.getcwd()
            # In ra thư mục hiện hành
            #print("Thư mục hiện hành là:", current_directory)
            #pdb.set_trace()
            self.read_xls()

#===============================================================================

    # Getter and Setter for adj
    @property
    def adj(self):
        return self._adj

    @adj.setter
    def adj(self, value):
        self._adj = value
        
    @property
    def M(self):
        return self._M
    @M.setter
    def M(self, value):
        self._M = value

    # Getter and Setter for H
    @property
    def H(self):
        return self._H

    @H.setter
    def H(self, value):
        self._H = value
        
    @property
    def draw(self):
        return self._draw
    @draw.setter
    def draw(self, value):
        self._draw = value

    # Getter and Setter for d
    @property
    def d(self):
        return self._d

    @d.setter
    def d(self, value):
        self._d = value

    # Getter and Setter for alpha
    @property
    def alpha(self):
        return self._alpha

    @alpha.setter
    def alpha(self, value):
        self._alpha = value

    # Getter and Setter for beta
    @property
    def beta(self):
        return self._beta

    @beta.setter
    def beta(self, value):
        self._beta = value

    # Getter and Setter for gamma
    @property
    def gamma(self):
        return self._gamma

    @gamma.setter
    def gamma(self, value):
        self._gamma = value

    # Getter and Setter for ID
    @property
    def ID(self):
        return self._ID

    @ID.setter
    def ID(self, value):
        self._ID = value

    # Getter and Setter for earliness
    @property
    def earliness(self):
        return self._earliness

    @earliness.setter
    def earliness(self, value):
        self._earliness = value    
    
    @property
    def target_nodes(self):
        return self._target_nodes
    
    @target_nodes.setter
    def target_nodes(self, value):
        self._target_nodes = value
        
    # Getter and Setter for tardiness
    @property
    def tardiness(self):
        return self._tardiness

    @tardiness.setter
    def tardiness(self, value):
        self._tardiness = value

    # Getter và Setter cho space_edges
    @property
    def space_edges(self):
        return self._space_edges

    @space_edges.setter
    def space_edges(self, value):
        if not isinstance(value, list):
            raise ValueError("space_edges must be a list")
        self._space_edges = value

    # Getter và Setter cho ts_edges
    @property
    def ts_edges(self):
        return self._ts_edges

    @ts_edges.setter
    def ts_edges(self, value):
        if not isinstance(value, list):
            raise ValueError("ts_edges must be a list")
        self._ts_edges = value

    # Getter và Setter cho ts_nodes
    @property
    def ts_nodes(self):
        return self._ts_nodes

    @ts_nodes.setter
    def ts_nodes(self, value):
        if not isinstance(value, list):
            raise ValueError("ts_nodes must be a list")
        self._ts_nodes = value

    # Getter và Setter cho tsedges
    @property
    def tsedges(self):
        return self._tsedges

    @tsedges.setter
    def tsedges(self, value):
        if not isinstance(value, list):
            raise ValueError("tsedges must be a list")
        self._tsedges = value

    # Getter và Setter cho started_nodes
    @property
    def started_nodes(self):
        return self._started_nodes

    @started_nodes.setter
    def started_nodes(self, value):
        if not isinstance(value, list):
            raise ValueError("started_nodes must be a list")
        self._started_nodes = value

    # Getter và Setter cho print_out
    @property
    def print_out(self):
        return self._print_out

    @print_out.setter
    def print_out(self, value):
        if not isinstance(value, bool):
            raise ValueError("print_out must be a boolean")
        self._print_out = value

    # Getter và Setter cho time_window_controller
    @property
    def time_window_controller(self):
        return self._time_window_controller

    @time_window_controller.setter
    def time_window_controller(self, value):
        self._time_window_controller = value

    # Getter và Setter cho restriction_controller
    @property
    def restriction_controller(self):
        return self._restriction_controller

    @restriction_controller.setter
    def restriction_controller(self, value):
        self._restriction_controller = value

    # Getter và Setter cho start_ban
    @property
    def start_ban(self):
        return self._start_ban

    @start_ban.setter
    def start_ban(self, value):
        if not isinstance(value, int):
            raise ValueError("start_ban must be an integer")
        self._start_ban = value

    # Getter và Setter cho end_ban
    @property
    def end_ban(self):
        return self._end_ban

    @end_ban.setter
    def end_ban(self, value):
        if not isinstance(value, int):
            raise ValueError("end_ban must be an integer")
        self._end_ban = value

    # Getter và Setter cho seed
    @property
    def seed(self):
        return self._seed

    @seed.setter
    def seed(self, value):
        if not isinstance(value, int):
            raise ValueError("seed must be an integer")
        self._seed = value

    # Getter và Setter cho num_max_agvs
    @property
    def num_max_agvs(self):
        return self._num_max_agvs

    @num_max_agvs.setter
    def num_max_agvs(self, value):
        self._num_max_agvs = value

    # Getter và Setter cho graph
    @property
    def graph(self):
        return self._graph

    @graph.setter
    def graph(self, value):
        self._graph = value
    
    def append_target(self, target_node):
        if isinstance(target_node, TimeWindowNode):
            #pdb.set_trace()
            pass
        self._target_nodes.append(target_node)
        
    def get_targets(self, index = -1):
        if (index != -1):
            return self._target_nodes[index]
        return self._target_nodes
    
    def get_target_by_id(self, id):
        for node in self._target_nodes:
            if(node.id == id):
                return node
        return None
    
    def process_number(self, num):
        import math
        if num < 5:
            return 0
        else:
            return math.ceil(num)
    
    def read_xls(self):
        #import pandas as pd
        from openpyxl import load_workbook
        import math
        
        # Đọc file Excel
        #pdb.set_trace()
        file_name = 'completion_times.xlsx'
        workbook = load_workbook(file_name, data_only=True)
        sheet = workbook.active
        
        # Find the last column with a value in the first row
        last_column_with_value = sheet.max_column
        for col in range(sheet.max_column, 0, -1):
            if sheet.cell(row=1, column=col).value is not None:
                last_column_with_value = col
                break
        # Lấy số lượng cột và hàng
        max_column = sheet.max_column
        max_row = sheet.max_row
        # Lấy dữ liệu từ 3 cột cuối cùng
        last_three_columns = []
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=last_column_with_value-2, max_col=last_column_with_value):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    #print(cell.value)
                    row_data.append(cell.value)
            if row_data:
                #print(row_data)
                last_three_columns.append(row_data)
        #df = pd.read_excel('completion_times.xlsx', engine='openpyxl')
        # Get the last 3 columns
        #last_three_columns = df.iloc[:, -3:]
        # Initialize an empty list to store the processed numbers
        self.processed_numbers = []
        
        for row in last_three_columns:
            for num in row:
                if isinstance(num, (int, float)):  # Kiểm tra nếu là số
                    self.processed_numbers.append(self.process_number(num))
                else:
                    #print(num)
                    pass
        

#======================================================================================

    def getReal(self, start_id, next_id, agv):
        M = self.graph.number_of_nodes_in_space_graph
        result = -1
        
        real_start_id, old_real_path = self._get_real_start_id_and_path(start_id, agv, M)
        start_time, end_time = self._calculate_times(start_id, next_id, M)
        space_start_node, space_end_node = self._get_space_nodes(start_id, next_id, M)
        
        min_moving_time = self._get_min_moving_time(space_start_node, space_end_node)
        end_time = max(end_time, start_time + min_moving_time)

        if self._is_target_node(next_id):
            result = 0
            self._update_agv_path(agv, next_id)
        
        result = self._handle_special_cases(start_id, next_id, start_time, end_time, result)
        
        if config.level_of_simulation == 2:
            result = self._calculate_sfm_runtime(space_start_node, space_end_node, agv, start_time, result)
        #elif config.level_of_simulation == 1 or config.level_of_simulation == 0:
        result = self._calculate_final_result(result, start_time, end_time)
        #else config.level_of_simulation == 0:
        #    pass
        
        
        
        return self._handle_collisions(result, next_id, agv, M)

    def _get_real_start_id_and_path(self, start_id, agv, M):
        if agv is None:
            return start_id % M + (M if start_id % M == 0 else 0), []
        old_real_path = [(node % M + (M if node % M == 0 else 0)) for node in agv.path]
        real_start_id = start_id % M + (M if start_id % M == 0 else 0)
        if real_start_id in old_real_path:
            return real_start_id, old_real_path
        agv.path.add(start_id)
        return real_start_id, old_real_path

    def _calculate_times(self, start_id, next_id, M):
        start_time = start_id // M - (1 if start_id % M == 0 else 0)
        end_time = next_id // M - (1 if next_id % M == 0 else 0)
        return start_time, end_time

    def _get_space_nodes(self, start_id, next_id, M):
        space_start_node = start_id % M + (M if start_id % M == 0 else 0)
        space_end_node = next_id % M + (M if next_id % M == 0 else 0)
        return space_start_node, space_end_node

    def _get_min_moving_time(self, space_start_node, space_end_node):
        edges_with_cost = {
            (int(edge[1]), int(edge[2])): [int(edge[4]), int(edge[5])]
            for edge in self.space_edges
            if edge[3] == '0' and int(edge[4]) >= 1
        }
        return edges_with_cost.get((space_start_node, space_end_node), [-1, -1])[1]

    def _is_target_node(self, next_id):
        all_ids_of_target_nodes = [node.id for node in self.target_nodes]
        return next_id in all_ids_of_target_nodes

    def _update_agv_path(self, agv, node_id):
        if agv is not None:
            agv.path.add(node_id)

    def _handle_special_cases(self, start_id, next_id, start_time, end_time, result):
        try:
            #if(next_id == 26306):
            #    pdb.set_trace()
            if isinstance(self.graph.nodes[next_id], TimeWindowNode):
                return end_time - start_time if result == -1 else result
        except KeyError:
            for e in self.ts_edges:
                if e[0] % self.graph.number_of_nodes_in_space_graph == start_id % self.graph.number_of_nodes_in_space_graph:
                    result = e[4] if result == -1 else result
            return abs(end_time - start_time) if result == -1 else result
        return result

    def _calculate_sfm_runtime(self, space_start_node, space_end_node, agv, start_time, result):
        runtime = self.getAGVRuntime(config.filepath, config.functions_file, space_start_node, space_end_node, agv, start_time)
        if runtime != -1:
            print(f"{bcolors.OKGREEN}{agv.id} from {space_start_node} to {space_end_node} at time {start_time} has runtime {runtime}.{bcolors.ENDC}")
            return runtime
        return result

    def _calculate_final_result(self, result, start_time, end_time):
        if result == -1:
            import random
            import time
            import math
            
            # Get the current time
            current_time = time.localtime()
            # Extract the components of the current time
            second = current_time.tm_sec
            minute = current_time.tm_min
            hour = current_time.tm_hour
            day = current_time.tm_mday
            month = current_time.tm_mon
            year = current_time.tm_year
            
            seed = second + minute * 60 + hour * 3600 + day * 86400 + month * 2592000 + year * 31104000
            
            # Seed the random number generator
            random.seed(seed)
            # Generate a random number with the given max value
            max_value = len(self.processed_numbers) if config.level_of_simulation == 1 else 300
            index = random.randint(0, max_value)
            if(config.level_of_simulation == 1):
                if index >= len(self.processed_numbers):
                    index = index % len(self.processed_numbers)
                    pdb.set_trace()
            value = self.processed_numbers[index] if config.level_of_simulation == 1 else index
            return math.ceil((1 + value/100)*(end_time - start_time))
            #return 3 if (end_time - start_time <= 3) else 2 * (end_time - start_time) - 3
        return result

    def _handle_collisions(self, result, next_id, agv, M):
        all_ids_of_target_nodes = [node.id for node in self.target_nodes]
        collision = True
        while collision:
            collision = False
            if next_id not in all_ids_of_target_nodes and next_id in self.graph.nodes:
                node = self.graph.nodes[next_id]
                if node.agv and node.agv != agv:
                    print(f'{node.agv.id} != {agv.id}')
                    collision = True
                    result += 1
                    next_id += M
        return result
    
    def getReal_preprocess(self, Map_file, function_file):
        # read files
        map_data = None
        function_data = None
        with open(Map_file, 'r', encoding='utf-8') as file:
            map_data = file.readlines()
        with open(function_file, 'r', encoding='utf-8') as file:
            function_data = file.readlines()
        hallways_list = []
        functions_list = []
        for line in map_data:
            line = line.strip()
            parts = line.split(" ")
            if len(parts) == 8:
                hallway = {
                    "hallway_id": parts[6],
                    "length": int(int(parts[5]) * 0.6),
                    "width": 4,
                    "agents_distribution": int(parts[7]),
                    "src": int(parts[1]),
                    "dest": int(parts[2])
                }
                hallways_list.append(hallway)
        for line in function_data:
            line = line.strip()
            functions_list.append(line)
        return hallways_list, functions_list

    def getAGVRuntime(self, Map_file, function_file, start_id, next_id, agv, current_time):
        hallways_list, functions_list = self.getReal_preprocess(Map_file, function_file)
        agv_id = self._extract_agv_id(agv)
        direction, hallway_id = self._get_hallway_direction(hallways_list, start_id, next_id)
        
        if hallway_id is None:
            print(f"{bcolors.WARNING}Hallway not found!{bcolors.ENDC}")
            return -1

        events_list = self._create_event_list(agv_id, direction, current_time, hallway_id)
        hallways_list = self._filter_hallways_list(hallways_list, hallway_id, direction)
        
        return self._simulate_bulk_runtime(agv_id, hallway_id, hallways_list, functions_list, events_list)

    def _extract_agv_id(self, agv):
        return int(agv.id[3:])

    def _get_hallway_direction(self, hallways_list, start_id, next_id):
        for hallway in hallways_list:
            if hallway["src"] == start_id and hallway["dest"] == next_id:
                return 1, hallway["hallway_id"]
            elif hallway["src"] == next_id and hallway["dest"] == start_id:
                return 0, hallway["hallway_id"]
        return 0, None

    def _create_event_list(self, agv_id, direction, time_stamp, hallway_id):
        event = {
            "AgvIDs": [agv_id],
            "AgvDirections": [direction],
            "time_stamp": int(time_stamp),
            "hallway_id": hallway_id
        }
        return [event]

    def _filter_hallways_list(self, hallways_list, hallway_id, direction):
        return [
            hallway for hallway in hallways_list
            if hallway["hallway_id"] == hallway_id and (hallway["src"] - hallway["dest"]) * direction > 0
        ]

    def _simulate_bulk_runtime(self, agv_id, hallway_id, hallways_list, functions_list, events_list):
        bulk_sim = BulkHallwaySimulator("test", 3600, hallways_list, functions_list, events_list)
        result = bulk_sim.run_simulation()
        completion_time = result[agv_id][hallway_id]["completion_time"]
        print(f"{bcolors.OKGREEN}AGV {agv_id} has runtime {completion_time} in hallway {hallway_id}.{bcolors.ENDC}")
        return completion_time
    
    def insertEdgesAndNodes(self, start, end, edge):
        start_id = start if isinstance(start, int) else start.id
        end_id = end if isinstance(end, int) else end.id
        self.graph.adjacency_list[start_id].append((end_id, edge))
        start_node = start if isinstance(start, Node) else self.find_node(start)
        end_node = end if isinstance(end, Node) else self.find_node(end)
        if self.graph.nodes[start_id] is None:
            self.graph.nodes[start_id] = start_node
        if self.graph.nodes[end_id] is None:
            self.graph.nodes[end_id] = end_node
            
    def find_unique_nodes(self, file_path = 'traces.txt'):
        """ Find nodes that are only listed as starting nodes in edges. """
        if not os.path.exists(file_path):
            print(f"File {file_path} does not exist.")
            return []
        
        target_ids = set()
        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                if line.startswith('a'):
                    parts = line.split()
                    target_ids.add(int(parts[3]))

        unique_ids = set()
        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                if line.startswith('a'):
                    parts = line.split()
                    node_id = int(parts[1])
                    if node_id not in target_ids:
                        unique_ids.add(node_id)

        return list(unique_ids)
    
    def update(self,currentpos,nextpos,realtime):
        list = utility()
        del self.matrix[currentpos,nextpos]
        q = deque()
        q.append(nextpos)
        while q:
            pos = q[0]
            q.pop()
            for i in list.findid(pos):
                if (pos,i) in self.matrix:
                    del self.matrix[pos,i]
                    q.append(i)
        nextpos = list.M*(int(currentpos/list.M)+ realtime) + list.getid(nextpos)
        self.matrix[currentpos,nextpos] = realtime
        q.append(nextpos)
        while q:
            pos = q[0]
            q.pop()
            for i in list.findid(pos):
                if (pos,i) not in self.matrix:
                    self.matrix[pos,i] = int((pos-i)/list.M)
                    q.append(i)
    
    def update_graph(self, id1=-1, id2=-1, end_id=-1, agv_id=None):
        """Cập nhật đồ thị với thông tin cạnh mới."""
        ID1, ID2, endid = self.get_ids(id1, id2, end_id)
        M = self.graph.number_of_nodes_in_space_graph
        current_time, new_node_id = self.calculate_times(ID1, ID2, endid, M)

        self.process_adjacency_list(current_time, new_node_id, M)
        
        q = self.update_new_started_nodes(new_node_id)
        new_edges = self.insert_from_queue(q, self.graph.adjacency_list)
        self.process_new_edges(new_edges)

        if self.version_check(current_time):
            self.graph.version += 1

        new_halting_edges = self.collect_new_halting_edges()
        self.graph.write_to_file([agv_id, new_node_id], new_halting_edges)                
    
    def process_adjacency_list(self, current_time, new_node_id, M):
        """Duyệt qua từng phần tử của adjacency_list và cập nhật thông tin."""
        for source_id, edges in list(self.graph.adjacency_list.items()):
            isContinued = any(node.id == source_id for node in self.target_nodes)
            if isContinued:
                continue

            if source_id in self.graph.nodes:
                node = self.graph.nodes[source_id]
                time = source_id // M - (1 if source_id % M == 0 else 0)
                if time < current_time and not isinstance(node, (TimeWindowNode, RestrictionNode)):
                    self.update_nodes(source_id, current_time, M)
    
    def update_nodes(self, source_id, current_time, M):
        """Cập nhật thông tin nút và xóa khỏi danh sách."""
        del self.graph.adjacency_list[source_id]
        if self.graph.nodes[source_id].agv is not None:
            space_id = M if (source_id % M == 0) else source_id % M
            new_source_id = current_time * M + space_id
            self.transfer_agv(source_id, new_source_id)
        del self.graph.nodes[source_id]
    
    def transfer_agv(self, source_id, new_source_id):
        """Chuyển AGV từ nút cũ sang nút mới."""
        try:
            if new_source_id in self.graph.nodes:
                self.graph.nodes[new_source_id].agv = self.graph.nodes[source_id].agv
            index = self.started_nodes.index(source_id)
            self.started_nodes[index] = new_source_id
        except ValueError:
            pass
        
    def get_ids(self, id1, id2, end_id):
        """Nhận ID từ người dùng hoặc sử dụng giá trị mặc định."""
        ID1 = int(input("Nhap ID1: ")) if id1 == -1 else id1
        ID2 = int(input("Nhap ID2: ")) if id2 == -1 else id2
        endid = int(input("Nhap ID thực sự khi AGV kết thúc hành trình: ")) if end_id == -1 else end_id
        return ID1, ID2, endid

    def calculate_times(self, ID1, ID2, endid, M):
        """Tính thời gian và ID nút mới."""
        time2 = ID1 // M - (1 if ID1 % M == 0 else 0)
        current_time = endid // M - (1 if endid % M == 0 else 0)
        new_node_id = current_time * M + (M if ID2 % M == 0 else ID2 % M)
        return current_time, new_node_id

    def update_new_started_nodes(self, new_node_id):
        """Cập nhật danh sách các nút mới bắt đầu và trả về hàng đợi."""
        q = deque([new_node_id])
        new_started_nodes = self.graph.getAllNewStartedNodes()
        for start in new_started_nodes:
            if start != new_node_id:
                q.append(start)
        return q

    def process_new_edges(self, new_edges):
        """Xử lý và cập nhật các cạnh mới vào đồ thị."""
        for edge in new_edges:
            #if edge == "a 599 1239 0 1 10": #or edge == "a 792 1432 0 1 10":#liệu đây có làm arr bị None???
            #    pdb.set_trace()
            arr = self.graph.parse_string(edge)
            if arr is not None:
                if arr[0] is None:
                    pdb.set_trace()
            else:
                pdb.set_trace()            
            source_id, dest_id = arr[0], arr[1]
            self.add_edge_to_graph(source_id, dest_id, arr)

    def add_edge_to_graph(self, source_id, dest_id, arr):
        """Thêm một cạnh mới vào đồ thị."""
        if source_id not in self.graph.nodes:
            self.graph.nodes[source_id] = self.find_node(source_id)
        if dest_id not in self.graph.nodes:
            self.graph.nodes[dest_id] = self.find_node(dest_id)

        if source_id not in self.graph.adjacency_list:
            self.graph.adjacency_list[source_id] = []
        
        found = any(end_id == dest_id for end_id, _ in self.graph.adjacency_list[source_id])
        if not found:
            anEdge = self.graph.nodes[source_id].create_edge(self.graph.nodes[dest_id], self.M, self.d, [source_id, dest_id, arr[2], arr[3], arr[4]])
            self.graph.adjacency_list[source_id].append([dest_id, anEdge])
        
        # Add TimeWindowEdge and RestrictionEdge
        self.time_window_controller.generate_time_window_edges(self.graph.nodes[source_id], self.graph.adjacency_list, self.graph.number_of_nodes_in_space_graph)
        self.restriction_controller.generate_restriction_edges(self.graph.nodes[source_id], self.graph.nodes[dest_id], self.graph.nodes, self.graph.adjacency_list)

    def version_check(self, current_time):
        """Kiểm tra nếu phiên bản cần được cập nhật."""
        time2 = self.graph.number_of_nodes_in_space_graph // self.M - (1 if self.graph.number_of_nodes_in_space_graph % self.M == 0 else 0)
        return time2 != current_time

    def collect_new_halting_edges(self):
        """Thu thập các cạnh dừng mới cần được thêm vào."""
        sorted_edges = sorted(self.graph.adjacency_list.items(), key=lambda x: x[0])
        new_nodes = set()
        new_halting_edges = []

        for source_id, edges in sorted_edges:
            for edge in edges:
                t = edge[0] // self.M - (1 if edge[0] % self.M == 0 else 0)
                if t >= self.H and edge[0] not in new_nodes and isinstance(self.graph.nodes[edge[0]], TimeoutNode):
                    new_nodes.add(edge[0])
                    for target in self.get_targets():
                        dest_id = target.id
                        new_halting_edges.append([edge[0], dest_id, 0, 1, self.H * self.H])

        return new_halting_edges
    
    def reset_agv(self, real_node_id, agv):
        for node_id in self.graph.nodes.keys():
            if(node_id != real_node_id):
                if self.graph.nodes[node_id].agv == agv:
                    self.graph.nodes[node_id].agv = None
        self.graph.nodes[real_node_id].agv = agv
    
    def remove_node_and_origins(self, node_id):
        node = None
        if isinstance(node_id, Node):
            node = node_id
        elif node_id in self.graph.nodes:
            node = self.graph.nodes[node_id]
        else:
            return
        node = node_id if isinstance(node_id, Node) else self.graph.nodes[node_id]
        R = [node]  # Khởi tạo danh sách R với nút cần xóa
        while R:  # Tiếp tục cho đến khi R rỗng
            current_node = R.pop()  # Lấy ra nút cuối cùng từ R
            if current_node.id in self.graph.nodes:  # Kiểm tra xem nút có tồn tại trong đồ thị hay không
                del self.graph.nodes[current_node.id]  # Nếu có, xóa nút khỏi danh sách các nút
            for id in self.graph.adjacency_list:
                edges = []
                found = False
                for end_id, edge in self.graph.adjacency_list[id]:
                    if(end_id == node.id):
                        #del self.adjacency_list
                        found = True
                    else:
                        edges.append([end_id, edge])
                if(found):
                    self.graph.adjacency_list[id] = edges
            
    def remove_edge(self, start_node, end_node, agv_id):
        if (start_node, end_node) in self.graph.edges:
            del self.graph.edges[(start_node, end_node)]
            self.graph.lastChangedByAGV = agv_id  # Update the last modified by AGV

    def handle_edge_modifications(self, start_node, end_node, agv):
        # Example logic to adjust the weights of adjacent edges
        print(f"Handling modifications for edges connected to {start_node} and {end_node}.")
        #pdb.set_trace()
        adjacent_nodes_with_weights = self.graph.adjacency_list.get(end_node, [])
        # Check adjacent nodes and update as necessary
        for adj_node, weight in adjacent_nodes_with_weights:
            if (end_node, adj_node) not in self.graph.lastChangedByAGV or self.graph.lastChangedByAGV[(end_node, adj_node)] != agv.id:
                # For example, increase weight by 10% as a traffic delay simulation
                new_weight = int(weight * 1.1)
                self.graph.adjacency_list[end_node][adj_node] = new_weight
                print(f"Updated weight of edge {end_node} to {adj_node} to {new_weight} due to changes at {start_node}.")
 
    def process_input_file(self, filepath):
        self.space_edges = []
        try:
            with open(filepath, 'r') as file:
                self.M = 0
                for line in file:
                    parts = line.strip().split()
                    if parts[0] == 'a' and len(parts) >= 6:
                        id1, id2 = int(parts[1]), int(parts[2])
                        self.space_edges.append(parts)
                        self.M = max(self.M, id1, id2)
                    elif parts[0] == 'n':
                        if(parts[2] == '1'):
                            self.started_nodes.append(int(parts[1]))
                        if(parts[2] == '-1'):
                            self.ID.append(int(parts[1]))
                            if isinstance(self.earliness, int):
                                self.earliness = []
                            if isinstance(self.tardiness, int):
                                self.tardiness = []
                            self.earliness.append(int(parts[3]))
                            self.tardiness.append(int(parts[4]))
                    elif parts[0] == 'alpha':
                        self.alpha = int(parts[1])
                    elif parts[0] == 'beta':
                        self.beta = int(parts[1])
            config.M = self.M
            if(self.print_out):
                print("Doc file hoan tat, M =", self.M)
        except FileNotFoundError:
            if(self.print_out):
                print("File khong ton tai!")
            return

    def find_node(self, _id):
        _id = int(_id)
        # Tìm kiếm đối tượng Node có ID tương ứng
        if not hasattr(self, 'map_nodes'):
            # Nếu chưa tồn tại, chuyển self.ts_nodes thành self.map_nodes
            self.map_nodes = {node.id: node for node in self.ts_nodes}
        # Tìm kiếm trên self.map_nodes
        if _id in self.map_nodes:
            return self.map_nodes[_id]
        else:
            # Nếu không có trên map_nodes, thêm vào cả ts_nodes và map_nodes
            #if(id == 26272):
            #    pdb.set_trace()
            for node in self._target_nodes:
                if(node.id == _id):
                    self.map_nodes[_id] = node
                    return node
            time = _id // self.M - (1 if _id % self.M == 0 else 0)
            new_node = None
            if(time >= self.H):
                new_node = TimeoutNode(_id, "TimeOut")
            else:
                new_node = Node(_id)
            self.ts_nodes.append(new_node)
            self.map_nodes[_id] = new_node
            
            return new_node
	
    def generate_hm_matrix(self):
        self.matrix = [[j + 1 + self.M * i for j in range(self.M)] for i in range(self.H)]
        if(self.print_out):
            print("Hoan tat khoi tao matrix HM!")
        #     print(' '.join(map(str, row)))

    def generate_adj_matrix(self):
        size = (self.H + 1) * self.M + 1
        self.adj = lil_matrix((2*size, 2*size), dtype=int)

        for edge in self.space_edges:
            if len(edge) >= 6 and edge[3] == '0' and int(edge[4]) >= 1:
                u, v, c = int(edge[1]), int(edge[2]), int(edge[5])
                for i in range(self.H + 1):
                    source_idx = i * self.M + u
                    target_idx = (i + c) * self.M + v
                    if(self.print_out):
                        print(f"i = {i} {source_idx} {target_idx} = 1")

                    if source_idx < size and (target_idx < size or (size <= target_idx < 2*size)):
                        self.adj[source_idx, target_idx] = 1
                        
        for i in range(size):
            j = i + self.M * self.d
            if j < size and (i % self.M == j % self.M):
                self.adj[i, j] = 1

        if(self.print_out):
            print("Hoan tat khoi tao Adjacency matrix!")

        rows, cols = self.adj.nonzero()
        with open('adj_matrix.txt', 'w') as file:
            for i, j in zip(rows, cols):
                file.write(f"({i}, {j})\n")
        if(self.print_out):
            print("Cac cap chi so (i,j) khac 0 cua Adjacency matrix duoc luu tai adj_matrix.txt.")

    def check_and_add_nodes(self, args, is_artificial_node = False, label = ""):
        if not hasattr(self, 'map_nodes'):
            # Nếu chưa tồn tại, chuyển self.ts_nodes thành self.map_nodes
            self.map_nodes = {node.id: node for node in self.ts_nodes}
        for id in args:
            # Ensure that Node objects for id exist in ts_nodes
            if not any(node.id == id for node in self.ts_nodes) and isinstance(id, int):
                if(is_artificial_node):
                   if(label == "TimeWindow"):
                       temp = TimeWindowNode(id, label)
                       self.ts_nodes.append(temp)
                       self.map_nodes[id] = temp
                   elif(label == "Restriction"):
                       temp = RestrictionNode(id, label)
                       self.ts_nodes.append(temp)
                       self.map_nodes[id] = temp
                   elif (label == "Timeout"):
                       temp = TimeoutNode(id, label)
                       self.ts_node.append(temp)
                       self.map_nodes[id] = temp
                   else:
                       temp = ArtificialNode(id, label)
                       self.ts_nodes.append(temp)
                       self.map_nodes[id] = temp
                else:
                    time = id // self.M - (1 if id % self.M == 0 else 0)
                    temp = None
                    if(time >= self.H):
                        temp = TimeoutNode(id, "Timeout")
                    else:
                        temp = Node(id)
                    self.ts_nodes.append(temp)
                    self.map_nodes[id] = temp
        #    self.ts_nodes.append(Node(ID2))

    def show(self, q):
        if len(q) < 10:
            return list(q)
        else:
            return list(q)[:5] + ["..."]

    def insert_from_queue(self, q, checking_list=None):
        """Chèn các cạnh từ hàng đợi vào đồ thị."""
        output_lines = []
        edges_with_cost = self.extract_edges_with_cost()
        ts_edges = self.get_ts_edges(checking_list)
        
        count = 0
        while q:
            if count % 1000 == 0:
                pass  # Có thể thêm log tại đây nếu cần
            count += 1
            ID = q.popleft()
            
            if not self.is_valid_id(ID):
                continue
            
            for j in self.adj.rows[ID]:
                if self.is_edge_present(ID, j, ts_edges):
                    continue
                
                self.add_edge_to_queue(q, ID, j, output_lines, edges_with_cost, checking_list)

        if checking_list is None:
            self.validate_edges()
        return output_lines

    def extract_edges_with_cost(self):
        """Trích xuất các cạnh có chi phí từ danh sách cạnh không gian."""
        return {(int(edge[1]), int(edge[2])): [int(edge[4]), int(edge[5])] 
                for edge in self.space_edges if edge[3] == '0' and int(edge[4]) >= 1}

    def get_ts_edges(self, checking_list):
        """Lấy danh sách các cạnh tạm thời."""
        if checking_list is None:
            return self.ts_edges
        return [[item[1].start_node.id, item[1].end_node.id] 
                for sublist in checking_list.values() for item in sublist]

    def is_valid_id(self, ID):
        """Kiểm tra xem ID có hợp lệ không."""
        return 0 <= ID < self.adj.shape[0]

    def is_edge_present(self, ID, j, ts_edges):
        """Kiểm tra xem cạnh đã tồn tại trong ts_edges chưa."""
        return any(edge[0] == ID and edge[1] == j for edge in ts_edges)

    def add_edge_to_queue(self, q, ID, j, output_lines, edges_with_cost, checking_list):
        """Thêm cạnh vào hàng đợi và ghi lại vào output_lines nếu cần."""
        if j not in q:
            q.append(j)
        
        u, v = self.get_node_coordinates(ID, j)
        cost_info = edges_with_cost.get((u, v), (-1, -1))
        
        if self.should_add_edge(cost_info, ID, j, v):
            self.create_edge_output(output_lines, ID, j, cost_info, checking_list)
        elif ID + self.M * self.d == j and ID % self.M == j % self.M:
            self.create_holding_edge_output(output_lines, ID, j, checking_list)

    def get_node_coordinates(self, ID, j):
        """Lấy tọa độ nút từ ID."""
        u = ID % self.M if ID % self.M != 0 or ID == 0 else self.M
        v = j % self.M if j % self.M != 0 or j == 0 else self.M
        return u, v

    def should_add_edge(self, cost_info, ID, j, v):
        """Kiểm tra điều kiện để thêm cạnh."""
        upper, cost = cost_info
        return ((ID // self.M) + cost >= (j // self.M) - (v // self.M)) and (upper != -1)

    def create_edge_output(self, output_lines, ID, j, cost_info, checking_list):
        """Tạo dòng output cho cạnh mới và thêm vào danh sách."""
        upper, cost = cost_info
        if ID // self.M >= self.H:
            output_lines.append(f"a {ID} {j} 0 1 {cost} Exceed")
        else:
            output_lines.append(f"a {ID} {j} 0 {upper} {cost}")
        
        if checking_list is None:
            self.ts_edges.append((ID, j, 0, upper, cost))
        
        self.check_and_add_nodes([ID, j])
        edge = self.find_node(ID).create_edge(self.find_node(j), self.M, self.d, [ID, j, 0, upper, cost])
        if checking_list is None:
            self.tsedges.append(edge)

    def create_holding_edge_output(self, output_lines, ID, j, checking_list):
        """Tạo dòng output cho cạnh holding và thêm vào danh sách."""
        output_lines.append(f"a {ID} {j} 0 1 {self.d}")
        
        if checking_list is None:
            self.ts_edges.append((ID, j, 0, 1, self.d))
        
        self.check_and_add_nodes([ID, j])
        edge = self.find_node(ID).create_edge(self.find_node(j), self.M, self.d, [ID, j, 0, 1, self.d])
        if checking_list is None:
            self.tsedges.append(edge)

    def validate_edges(self):
        """Kiểm tra tính nhất quán của các cạnh."""
        assert len(self.ts_edges) == len(self.tsedges), f"Thiếu cạnh ở đâu đó rồi {len(self.tsedges)} != {len(self.ts_edges)}"

    def create_tsg_file(self):          
        #pdb.set_trace()
        q = deque()
        q.extend(self.started_nodes)

        #pdb.set_trace()
        output_lines = self.insert_from_queue(q)
        with open('TSG.txt', 'w') as file:
            for line in output_lines:
                file.write(line + "\n")
        if(self.print_out):
            print("TSG.txt file created.")

    def init_agvs_n_events(self, all_agvs, events, graph, graph_processor):
        from controller.EventGenerator import StartEvent
        StartEvent.static_index = 0
        from model.AGV import AGV
        for node_id in self.started_nodes:
            #pdb.set_trace()
            agv = AGV("AGV" + str(node_id), node_id, graph)  # Create an AGV at this node
            #print(Event.getValue("number_of_nodes_in_space_graph"))
            if(self.M == 0):
                pdb.set_trace()
            start_time = node_id // self.M
            end_time = start_time
            start_event = StartEvent(start_time, end_time, agv, graph, graph_processor)  # Start event at time 0
            events.append(start_event)
            all_agvs.add(agv)  # Thêm vào tập hợp AGV
    
    def init_tasks(self, tasks):
        for node_id in self.get_targets():
            tasks.add(node_id)
    
    def query_edges_by_source_id(self):
        source_id = int(input("Nhap vao ID nguon: "))

        edges = []
        try:
            with open('TSG.txt', 'r') as file:
                for line in file:
                    parts = line.strip().split()
                    if parts[0] == 'a' and int(parts[1]) == source_id:
                        edges.append(line.strip())
        except FileNotFoundError:
            if(self.print_out):
                print("File TSG.txt khong ton tai!")
            return

        if edges:
            if(self.print_out):
                print(f"Cac canh co ID nguon la {source_id}:")
            for edge in edges:
                print(edge)
        else:
            if(self.print_out):
                print(f"Khong tim thay canh nao co ID nguon la {source_id}.")

    def init_nodes_n_edges(self):
        for edge in self.tsedges:
            if edge is not None:
                self.insertEdgesAndNodes(edge.start_node, edge.end_node, edge)
    
    def check_file_conditions(self):
        try:
            seen_edges = set()
            with open('TSG.txt', 'r') as file:
                for line in file:
                    parts = line.strip().split()
                    if parts[0] != 'a':
                        continue
                    ID1, ID2 = int(parts[1]), int(parts[2])

                    # Condition 1: ID1 should not equal ID2
                    if ID1 == ID2:
                        print("False")
                        return

                    # Condition 2: If ID1 before ID2, then ID2 should not come before ID1
                    if (ID1, ID2) in seen_edges or (ID2, ID1) in seen_edges:
                        print("False")
                        return
                    else:
                        seen_edges.add((ID1, ID2))

                    # Condition 3: ID2/self.M should be greater than ID1/self.M
                    if ID2 // self.M <= ID1 // self.M:
                        print("False")
                        return

            print("True")
        except FileNotFoundError:
            print("File TSG.txt khong ton tai!")

    def update_file(self, id1=-1, id2=-1, c12=-1):
        """Cập nhật file TSG.txt với các cạnh mới dựa trên đầu vào."""
        ID1 = self.get_input_id(id1, "Nhap ID1: ")
        ID2 = self.get_input_id(id2, "Nhap ID2: ")
        C12 = self.get_input_weight(c12)

        ID2 = self.adjust_id2_if_needed(ID1, ID2, C12)

        existing_edges = self.load_existing_edges()
        if (ID1, ID2) not in existing_edges:
            new_edges = self.find_new_edges(ID1, ID2, C12)
            self.append_edges_to_file(new_edges)

    def get_input_id(self, default_id, prompt):
        """Lấy ID từ người dùng hoặc sử dụng giá trị mặc định."""
        return int(input(prompt)) if default_id == -1 else default_id

    def get_input_weight(self, default_weight):
        """Lấy trọng số từ người dùng hoặc sử dụng giá trị mặc định."""
        return int(input("Nhap trong so C12: ")) if default_weight == -1 else default_weight

    def adjust_id2_if_needed(self, ID1, ID2, C12):
        """Điều chỉnh ID2 nếu i2 - i1 không bằng C12."""
        i1, i2 = ID1 // self.M, ID2 // self.M
        if i2 - i1 != C12:
            print('Status: i2 - i1 != C12')
            ID2 = ID1 + self.M * C12
        return ID2

    def load_existing_edges(self):
        """Tải các cạnh đã tồn tại từ file TSG.txt."""
        existing_edges = set()
        try:
            with open('TSG.txt', 'r') as file:
                for line in file:
                    parts = line.strip().split()
                    if parts[0] == 'a' and len(parts) >= 3 and parts[1].isdigit() and parts[2].isdigit():
                        existing_edges.add((int(parts[1]), int(parts[2])))
        except FileNotFoundError:
            print("File TSG.txt khong ton tai!")
            return existing_edges
        return existing_edges

    def find_new_edges(self, ID1, ID2, C12):
        """Tìm các cạnh mới cần thêm vào đồ thị."""
        q = deque([ID2])
        visited = {ID2}
        new_edges = [(ID1, ID2, C12)]

        while q:
            ID = q.popleft()
            for j in self.adj.rows[ID]:
                if j not in visited:
                    c = self.d if ID + self.M * self.d == j and ID % self.M == j % self.M else C12
                    if (ID // self.M) + c == j // self.M:
                        new_edges.append((ID, j, c))
                        q.append(j)
                        visited.add(j)
        
        return new_edges

    def append_edges_to_file(self, new_edges):
        """Thêm các cạnh mới vào file TSG.txt."""
        edges_with_cost = { (int(edge[1]), int(edge[2])): [int(edge[4]), int(edge[5])] 
                            for edge in self.space_edges if edge[3] == '0' and int(edge[4]) >= 1 }

        with open('TSG.txt', 'a') as file:
            for ID, j, c in new_edges:
                u, v = ID % self.M + (self.M if ID % self.M == 0 else 0), j % self.M + (self.M if j % self.M == 0 else 0)
                [upper, _] = edges_with_cost[(u, v)]
                file.write(f"a {ID} {j} 0 {upper} {c}\n")
        print("Da cap nhat file TSG.txt.")

    def add_restrictions(self):
        alpha = input("Nhập vào alpha: ")
        beta = input("Nhập vào beta: ")
        gamma = input("Nhập vào gamma: ")
        self.alpha = int(alpha) if alpha else 1
        self.beta = int(beta) if beta else 1
        self.gamma = int(gamma) if gamma else 1
        restriction_count = input("Hãy nhập số lượng các restriction: ")
        self.restriction_count = int(restriction_count) if restriction_count else 1
        start_ban, end_ban = map(int, input("Khung thời gian cấm (nhập hai số phân tách bằng khoảng trắng a b): ").split())
        self.start_ban = start_ban
        self.end_ban = end_ban
        self.restrictions = []

        for i in range(self.restriction_count):
            print(f"Restriction {i + 1}:")
            u, v = map(int, input("\tKhu vực cấm (nhập hai số phân tách bằng khoảng trắng a b): ").split())

            self.restrictions.append((u, v))
        self.ur = int(input("Số lượng hạn chế: "))

    def create_set_of_edges(self, edges):
        for e in edges:
            #self.tsedges.append(ArtificialEdge(self.find_node(e[0]), self.find_node(e[1]), e[4]))
            temp = self.find_node(e[0]).create_edge(self.find_node(e[1]), self.M, self.d, e)
            self.tsedges.append(temp)
        
    def process_restrictions(self):
        """Xử lý các hạn chế trong đồ thị."""
        if self.restriction_controller is None:
            self.restriction_controller = RestrictionController(self)

        edges_with_cost = self.get_edges_with_cost()
        maxid = self.get_max_id() + 1
        new_a = set()

        for restriction in self.restrictions:
            R = self.create_restricted_edges(restriction, edges_with_cost, maxid)
            if R:
                new_a.update(self.create_new_edges(restriction, R, maxid))
                maxid += 3

        self.update_edges(new_a)
        self.insert_halting_edges()
        self.write_to_file()

    def get_edges_with_cost(self):
        """Trả về một từ điển các cạnh với chi phí."""
        return {(int(edge[1]), int(edge[2])): int(edge[5])
                for edge in self.space_edges if edge[3] == '0' and int(edge[4]) >= 1}

    def create_restricted_edges(self, restriction, edges_with_cost, maxid):
        """Tạo các cạnh bị cấm dựa trên hạn chế và chi phí."""
        R = []
        for time in range(self.start_ban, self.end_ban + 1):
            time_space_point_0 = time * self.M + restriction[0]
            cost = edges_with_cost.get((restriction[0], restriction[1]), -1)
            time_space_point_1 = (time + cost) * self.M + restriction[1]
            
            R.append([time_space_point_0, time_space_point_1, cost])
            self.adj[time_space_point_0, time_space_point_1] = 0

        self.update_edges_after_restrictions(R)
        return R

    def update_edges_after_restrictions(self, R):
        """Cập nhật danh sách các cạnh sau khi áp dụng hạn chế."""
        assert len(self._edges) == len(self.tsedges), f"Thiếu cạnh ở đâu đó rồi {len(self.ts_edges)} != {len(self.ts_edges)}"
        self.ts_edges = [e for e in self._edges if [e[0], e[1]] not in [r[:2] for r in R]]
        self.tsedges = [e for e in self.tsedges if [e.start_node.id, e.end_node.id] not in [r[:2] for r in R]]

    def create_new_edges(self, restriction, R, maxid):
        """Tạo các cạnh mới dựa trên các hạn chế đã chỉ định."""
        a_s, a_t, a_sub_t = maxid, maxid + 1, maxid + 2
        self.check_and_add_nodes([a_s, a_t, a_sub_t], True, "Restriction")

        self.restriction_controller.add_nodes_and__re_node(
            R[0][0], R[0][1], restriction, a_s, a_t
        )

        new_edges = {
            (a_s, a_t, 0, self.H, int(self.gamma/self.alpha)),
            (a_s, a_sub_t, 0, self.ur, 0),
            (a_sub_t, a_t, 0, self.H, 0)
        }

        for e in R:
            new_edges.add((e[0], a_s, 0, 1, 0))
            new_edges.add((a_t, e[1], 0, 1, e[2]))

        return new_edges

    def update_edges(self, new_a):
        """Cập nhật danh sách các cạnh với các cạnh mới và đảm bảo tính chính xác."""
        self.ts_edges.extend(e for e in new_a if e not in self.ts_edges)
        self.create_set_of_edges(new_a)
        assert len(self.ts_edges) == len(self.tsedges), f"Thiếu cạnh ở đâu đó rồi {len(self.ts_edges)} != {len(self.tsedges)}"
        self.ts_edges.sort(key=lambda edge: (edge[0], edge[1]))
        
    def insert_halting_edges(self):
        halting_nodes = set()
        for edge in self.tsedges:
            if(isinstance(edge.end_node, TimeWindowNode)):
                continue
            time = edge.end_node.id // self.M - (1 if edge.end_node.id % self.M == 0 else 0)
            #    pdb.set_trace()
            if(time >= self.H):
                halting_nodes.add(edge.end_node.id)
        targets = self.get_targets()
        new_a = set()
        for h_node in halting_nodes:
            for target in targets:
                e = (h_node, target.id, 0, 1, self.H*self.H)
                new_a.update({e})
        self.ts_edges.extend(e for e in new_a if e not in self.ts_edges)
        self.create_set_of_edges(new_a)
    
    def write_to_file(self):
        M = max(target.id for target in self.get_targets())
        with open('TSG.txt', 'w') as file:
            file.write(f"p min {M} {len(self.ts_edges)}\n")
            for start in self.started_nodes:
                file.write(f"n {start} 1\n")
            for target in self.get_targets():
                target_id = target.id
                file.write(f"n {target_id} -1\n")
            #for edge in self.ts_edges:
            for edge in self.tsedges:
                if (edge is not None):   
                    if(edge.weight == self.H*self.H):
                        #pdb.set_trace()
                        file.write(f"c Exceed {edge.weight} {edge.weight // self.M}\na {edge.start_node.id} {edge.end_node.id} {edge.lower} {edge.upper} {edge.weight}\n")
                    else:
                        file.write(f"a {edge.start_node.id} {edge.end_node.id} {edge.lower} {edge.upper} {edge.weight}\n")
        if(self.print_out):
            print("Đã cập nhật các cung mới vào file TSG.txt.")
        
    def get_started_points(self):
        N = int(input("Nhập vào số lượng các xe AGV: "))
        self.started_nodes = []
        for i in range(1, N+1):
            p, t = map(int, input(f"Xe {i} xuất phát ở đâu và khi nào (nhập p t)?: ").split())
            p = t*self.M + p
            self.started_nodes.append(p)

    def get_max_id(self):
      max_val = 0
      try:
        with open('TSG.txt', 'r') as file:
            for line in file:
                parts = line.strip().split()
                if parts[0] == 'a':
                    max_val = max(max_val, int(parts[2]))
      except FileNotFoundError:
        pass
      return max_val
      
    def generate_numbers_student(self, G, H, M, N = 0, df=10):
        while True:
            self._seed = self._seed + 1
            self._seed = self._seed % G
            np.random.seed(self._seed)
            # Sinh 4 số ngẫu nhiên theo phân phối Student
            first_two = np.random.standard_t(df, size=2)
            numbers = np.random.standard_t(df, size=2)
            # Chuyển đổi các số này thành số nguyên trong khoảng từ 1 đến 100
            first_two = np.round((first_two - np.min(first_two)) / (np.max(first_two) - np.min(first_two)) * (G//3) + self._seed).astype(int)
            numbers = np.round((numbers - np.min(numbers)) / (np.max(numbers) - np.min(numbers)) * (H//3) + self._seed).astype(int)
            if first_two[0] < G and first_two[1] < G and numbers[0] <= numbers[1] and numbers[1] < H:
                # Kiểm tra điều kiện khoảng cách tối thiểu
                if (abs(first_two[0] - first_two[1]) >= M and abs(numbers[0] - numbers[1]) >= N):
                    return np.concatenate((first_two, numbers))
    
    def add_time_windows_constraints(self):
        from controller.TimeWindowController import TimeWindowController

        max_val = self.get_max_id() + 1
        target_node = self.create_time_window_node(max_val)

        if self.time_window_controller is None:
            self.time_window_controller = TimeWindowController(self.alpha, self.beta, self.gamma, self.d, self.H)

        ID, earliness, tardiness = self.get_initial_conditions(target_node)
        new_edges = self.process_tsg_file(target_node, ID, earliness, tardiness)

        self.update_edges(new_edges)

        if self.print_out:
            print(f"Đã cập nhật {len(new_edges)} cung mới vào file TSG.txt.")

    def create_time_window_node(self, max_val):
        target_node = TimeWindowNode(max_val, "TimeWindow")
        self.ts_nodes.append(target_node)
        self.append_target(target_node)
        return target_node

    def get_initial_conditions(self, target_node):
        if isinstance(self.ID, list):
            if(len(self.ID) == 0):
                pdb.set_trace()
            ID = self.ID[0]
            earliness = self.earliness[0]
            tardiness = self.tardiness[0]
            self.ID = self.ID[1:]
            self.earliness = self.earliness[1:]
            self.tardiness = self.tardiness[1:]
        else:
            ID = self.ID
            earliness = self.earliness
            tardiness = self.tardiness

        self.time_window_controller.add_source_and_TWNode(ID, target_node, earliness, tardiness)
        return ID, earliness, tardiness

    def process_tsg_file(self, target_node, ID, earliness, tardiness):
        new_edges = set()

        try:
            with open('TSG.txt', 'r') as file:
                for line in file:
                    parts = line.strip().split()
                    if parts[0] == 'a' and len(parts) >= 6:
                        ID2 = int(parts[2])
                        self.process_line(ID, ID2, earliness, tardiness, new_edges, target_node)

        except FileNotFoundError:
            pass

        return new_edges

    def process_line(self, ID, ID2, earliness, tardiness, new_edges, target_node):
        for i in range(1, self.H + 1):
            j = i * self.M + ID
            if j == ID2:
                C = int(int(self.beta) * max(earliness - i, 0, i - tardiness) / int(self.alpha))
                new_edges.add((j, target_node.id, 0, 1, C))
                self.find_node(j).create_edge(target_node, self.M, self.d, [j, target_node.id, 0, 1, C])
                break

        t = ID2 // self.M - (1 if ID2 % self.M == 0 else 0)
        if t > self.H:
            C = self.H * self.H
            new_edges.add((j, target_node.id, 0, 1, C))
            self.find_node(j).create_edge(target_node, self.M, self.d, [j, target_node.id, 0, 1, C])

    def update_edges(self, new_edges):
        self.ts_edges.extend(e for e in new_edges if e not in self.ts_edges)
        self.create_set_of_edges(new_edges)

        with open('TSG.txt', 'a') as file:
            for edge in new_edges:
                file.write(f"a {edge[0]} {edge[1]} {edge[2]} {edge[3]} {edge[4]}\n")

    def update_tsg_with_t(self):
        T = int(input("Nhập giá trị T: "))
        # Đảm bảo T là một giá trị nguyên dương
        if not isinstance(T, int) or T <= 0:
            print("Giá trị của T phải là một số nguyên dương.")
            return

        new_lines = []

        # Đọc và kiểm tra từng dòng trong file TSG.txt cũ
        try:
            with open('TSG.txt', 'r') as file:
                for line in file:
                    parts = line.strip().split()
                    if len(parts) == 6 and parts[0] == 'a':
                        ID1, ID2 = int(parts[1]), int(parts[2])

                        # Kiểm tra điều kiện ID1 và ID2
                        if ID1 >= T * self.M and ID2 >= T * self.M:
                            new_lines.append(line)
        except FileNotFoundError:
            print("Không tìm thấy file TSG.txt.")
            return

        # Ghi các dòng thỏa điều kiện vào file TSG.txt mới
        with open('TSG_new.txt', 'w') as file:
            for line in new_lines:
                file.write(line)
        print("Đã tạo file TSG_new.txt mới với các dòng thỏa điều kiện.")

    def add_problem_info(self):
      json_filepath = input("Nhap ten file dau vao: ")
      try:

          with open(json_filepath, 'r') as json_file:
              data = json.load(json_file)
              itinerary_start = data["itinerary"]["start"]
              itinerary_end = data["itinerary"]["end"]

              # Tính toán max_id và số lượng cung (A) từ file TSG.txt
              max_id = 0
              A = 0
              with open('TSG.txt', 'r') as tsg_file:
                  for line in tsg_file:
                      if line.startswith('a'):
                          A += 1
                          _, id1, id2, _, _, _ = line.split()
                          max_id = max(max_id, int(id1), int(id2))

              # Tạo dòng thông tin về bài toán cần giải
              problem_info_line = f"p min {max_id} {A}\n"

              # Tạo dòng thông tin về lịch trình
              itinerary_lines = []
              for item in itinerary_start:
                  time_values = item["time"]
                  for time_value in time_values:
                      point_id = item["point"] + self.M * time_value
                      itinerary_lines.append(f"n {point_id} 1\n")
              for item in itinerary_end:
                  point_id = item["point"][0]
                  time_values = item["time"]
                  itinerary_lines.append(f"n {point_id} -1\n")
                  self.ID = point_id
                  self.earliness = time_values[0]
                  self.tardiness = time_values[1]
                  self.alpha = 1
                  self.beta =  1
                  self.add_time_windows_constraints()

              # Ghi dòng thông tin về bài toán và lịch trình vào đầu file TSG.txt
              with open('TSG.txt', 'r+') as tsg_file:
                  content = tsg_file.read()
                  tsg_file.seek(0, 0)
                  tsg_file.write(problem_info_line + ''.join(itinerary_lines) + content)

              print("Đã thêm thông tin về bài toán vào file TSG.txt.")
      except FileNotFoundError:
          print("Không tìm thấy file JSON hoặc TSG.txt.")

    def remove_duplicate_lines(self):
            try:
                # Read lines from TSG.txt
                with open('TSG.txt', 'r') as file:
                    lines = file.readlines()

                seen_lines = set()
                unique_lines = []
                for line in lines:
                    if re.match(r'^a\s+\d+\s+\d+', line):
                        if line.strip() not in seen_lines:
                            unique_lines.append(line)
                            seen_lines.add(line.strip())
                    else:
                        unique_lines.append(line)

                # Write unique lines back to TSG.txt
                with open('TSG.txt', 'w') as file:
                    file.writelines(unique_lines)

                print("Removed duplicate lines from TSG.txt.")
            except FileNotFoundError:
                print("File TSG.txt not found.")

    def remove_redundant_edges(self):
        R, E, S = self.initialize_sets()
        
        if R is None or E is None:
            return  # Nếu gặp lỗi trong quá trình đọc file, thoát khỏi hàm.

        self.filter_edges(R, E)

    def initialize_sets(self):
        R = set()  # Tập ID của nút nguồn
        E = set()  # Tập ID của nút đích

        try:
            with open('TSG.txt', 'r') as file:
                lines = file.readlines()
                if not lines:
                    print("File rỗng.")
                    return None, None

                first_line = lines[0].strip()
                if first_line.startswith('p min'):
                    S = self.extract_node_ids(lines)

                    # Lưu ID của nút nguồn vào tập R
                    R = self.extract_source_ids(lines, S)

                    # Lưu ID của nút đích vào tập E
                    E = self.extract_target_ids(lines)

                else:
                    print("File không đúng định dạng.")
                    return None, None
        except FileNotFoundError:
            print("File không tồn tại.")
            return None, None

        return R, E

    def extract_node_ids(self, lines):
        S = set()
        for line in lines[1:]:
            if line.startswith('n'):
                _, node_id, _ = line.split()
                S.add(int(node_id))
        return S

    def extract_source_ids(self, lines, S):
        R = set()
        for line in lines[1:]:
            if line.startswith('a'):
                _, source_id, _, _, _, _ = line.split()
                source_id = int(source_id)
                if source_id not in S:
                    R.add(source_id)
        return R

    def extract_target_ids(self, lines):
        E = set()
        for line in lines[1:]:
            if line.startswith('a'):
                _, _, target_id, _, _, _ = line.split()
                E.add(int(target_id))
        return E

    def filter_edges(self, R, E):
        try:
            with open('TSG.txt', 'r') as file:
                lines = file.readlines()

            new_lines = []
            for line in lines:
                if line.startswith('a'):
                    _, source_id, target_id, _, _, _ = line.split()
                    source_id = int(source_id)
                    target_id = int(target_id)

                    # Nếu source_id không thuộc S và không thuộc E, loại bỏ cạnh
                    if source_id not in R and source_id not in E:
                        continue

                # Thêm dòng vào danh sách mới
                new_lines.append(line)

            # Ghi các dòng mới vào file TSG.txt
            with open('TSG.txt', 'w') as file:
                file.writelines(new_lines)

            print("Đã loại bỏ các cạnh dư thừa từ file TSG.txt.")
        except FileNotFoundError:
            print("File không tồn tại.")

    def remove_descendant_edges(self):
        source_id = int(input("Nhập ID của điểm gốc: "))
        try:
          with open('TSG.txt', 'r') as file:
              lines = file.readlines()
        except FileNotFoundError:
          print("File TSG.txt không tồn tại.")
          return

        # Tạo một hàng đợi để duyệt qua các cung
        queue = deque([source_id])

        # Dùng set để lưu trữ ID của các cung cần loại bỏ
        to_remove = set()

        while queue:
          current_id = queue.popleft()
          for line in lines:
              parts = line.strip().split()
              if len(parts) == 6 and parts[0] == 'a' and int(parts[1]) == current_id:
                  destination_id = int(parts[2])
                  to_remove.add(line.strip())
                  queue.append(destination_id)

        new_lines = [line for line in lines if line.strip() not in to_remove]
        # Ghi lại các cung còn lại vào file TSG.txt
        with open('TSG.txt', 'w') as file:
          file.writelines(new_lines)

        print("Đã gỡ bỏ các cung con cháu xuất phát từ điểm gốc trong đồ thị TSG.")
    
    def process_tsg(self):
        AGV, TASKS, objective_coeffs = self.initialize_sets()
        
        if AGV is None or objective_coeffs is None:
            return  # Nếu gặp lỗi trong quá trình đọc file, thoát khỏi hàm.

        solver = pywraplp.Solver.CreateSolver('SCIP')
        self.setup_objective(solver, AGV, objective_coeffs)

        status = solver.Solve()
        self.handle_solution(status, solver, AGV, objective_coeffs)

    def initialize_sets(self):
        AGV = set()
        TASKS = set()
        objective_coeffs = {}

        try:
            with open('TSG.txt', 'r') as file:
                for line in file:
                    parts = line.strip().split()
                    if len(parts) == 3 and parts[0] == 'n':
                        node_id, val = int(parts[1]), int(parts[2])
                        if val == 1:
                            AGV.add(node_id)
                        elif val == -1:
                            TASKS.add(node_id)
                    elif len(parts) == 6 and parts[0] == 'a':
                        i, j, c = int(parts[1]), int(parts[2]), int(parts[5])
                        if (i, j) not in objective_coeffs:
                            objective_coeffs[(i, j)] = c

        except FileNotFoundError:
            print("File TSG.txt không tồn tại.")
            return None, None, None

        return AGV, TASKS, objective_coeffs

    def setup_objective(self, solver, AGV, objective_coeffs):
        objective = solver.Objective()
        added_keys = set()  # Sử dụng set để lưu trữ các khóa đã được thêm
        
        for m in AGV:
            for i, j in objective_coeffs.keys():
                key = f'x_{m}_{i}_{j}'
                if key not in added_keys:  # Chỉ tạo biến nếu chưa được tạo trước đó
                    x = solver.BoolVar(key)
                    objective.SetCoefficient(x, objective_coeffs[(i, j)])  # Đặt hệ số cho mỗi biến
                    added_keys.add(key)  # Thêm khóa vào set đã được thêm

        objective.SetMinimization()
        print(added_keys)

        # Thêm ràng buộc x_m_i_j nhận giá trị 0 hoặc 1
        self.add_constraints(solver, AGV, objective_coeffs)

    def add_constraints(self, solver, AGV, objective_coeffs):
        for m in AGV:
            for i, j in objective_coeffs.keys():
                x = solver.LookupVariable(f'x_{m}_{i}_{j}')
                constraint = solver.Constraint(0, 1)
                constraint.SetCoefficient(x, 1)

    def handle_solution(self, status, solver, AGV, objective_coeffs):
        if status == pywraplp.Solver.OPTIMAL:
            print('Solution:')
            print('Objective value =', solver.Objective().Value())
            for m in AGV:
                for i, j in objective_coeffs.keys():
                    x = solver.LookupVariable(f'x_{m}_{i}_{j}')
                    if x.solution_value() == 1:
                        print(f'x_{m}_{i}_{j} = 1')
        else:
            print('The problem does not have an optimal solution.')
     
    def generate_poisson_random(self, M = None):
        if M is None:
            M = self.M
        if M <= 2 and M >= 1:
            return M
        while True:
            # Sinh số ngẫu nhiên theo phân phối Poisson
            number = np.random.poisson(lam=M)        
            # Kiểm tra điều kiện số ngẫu nhiên lớn hơn 1 và nhỏ hơn hoặc bằng M
            if 1 < number < M:
                return number

    def use_in_main(self, use_config_data = False, print_output = False):
        self.print_out = print_output
        if(use_config_data):
            filepath = config.filepath
        else:
            filepath = input("Nhap ten file can thuc hien (hint: Redundant3x3Wards.txt): ")
            if filepath == '':
                filepath = 'Redundant3x3Wards.txt'
            config.filepath = filepath
        self.started_nodes = [] #[1, 10]

        self.process_input_file(filepath)
        if(use_config_data):
            self.H = config.H
        else:
            self.H = input("Nhap thoi gian can gia lap (default: 10): ")
            if(self.H == ''):
                self.H = 10
            else:
                self.H = int(self.H)
            config.H = self.H
        
        if(use_config_data):
            self.draw = config.draw
        else:
            self.draw = input("Nhập 1 để vẽ TSG (default 0) không nên dùng với đồ thị lớn): ")
            if(self.draw == '' or self.draw == 0):
                self.draw = 0
            else:
                self.draw = 1
            config.draw = self.draw

        self.generate_hm_matrix()
        if(use_config_data):
            self.d = config.d
        else:
            self.d = input("Nhap time unit (default: 10): ")
            if(self.d == ''):
                self.d = 10
            else:
                self.d = int(self.d)
            config.d = self.d
        
        self.generate_adj_matrix()
        
        num_of_agvs = 0
        if(use_config_data):
            #pdb.set_trace()
            self.num_max_agvs = config.num_max_agvs
            self.ID = config.ID
            self.earliness = config.earliness
            self.tardiness = config.tardiness
            for i in range(len(config.started_nodes)):
                if (config.started_nodes[i] > self.M and config.started_nodes[i] % self.M not in config.started_nodes):
                    config.started_nodes[i] = config.started_nodes[i] % self.M
            self.started_nodes = config.started_nodes
            num_of_agvs = config.numOfAGVs
            if(config.numOfAGVs > len(config.ID)):
                num_of_additional_agvs = config.numOfAGVs - len(config.ID)
                for _ in range(num_of_additional_agvs):
                    [s, d, e, t] = self.generate_numbers_student(self.M, self.H, 12, 100)
                    while(d in self.ID or s in self.started_nodes):
                        [s, d, e, t] = self.generate_numbers_student(self.M, self.H, 12, 100)
                    self.started_nodes.append(s)
                    self.ID.append(d)
                    config.ID.append(d)
                    self.earliness.append(e)
                    self.tardiness.append(t)
            elif(config.numOfAGVs < len(config.ID)):
                config.ID = config.ID[:(config.numOfAGVs)]
                config.earliness = config.earliness[:(config.numOfAGVs)]
                config.tardiness = config.tardiness[:(config.numOfAGVs)]
                config.started_nodes = config.started_nodes[:(config.numOfAGVs)]
                self.earliness = self.earliness[:(config.numOfAGVs)]
                self.tardiness = self.tardiness[:(config.numOfAGVs)]
                self.started_nodes = self.started_nodes[:(config.numOfAGVs)]
        else:
            self.num_max_agvs = input("Nhap so luong AGV toi da di chuyen trong toan moi truong (default: 2):")
            if(self.num_max_agvs == ''):
                self.num_max_agvs = 2
            else:
                self.num_max_agvs = int(self.num_max_agvs)
            num_of_agvs = self.num_max_agvs
            config.num_max_agvs = self.num_max_agvs
            config.numOfAGVs = num_of_agvs
            if len(self.started_nodes) == 0:
                self.ID = []
                self.earliness = []
                self.tardiness = []
                #pdb.set_trace()
                for _ in range(num_of_agvs):
                    [s, d, e, t] = self.generate_numbers_student(self.M, self.H, int(0.2*self.M))#, 100 if self.H > 100 else self.H//3)
                    while s in self.started_nodes:
                        s += self.M
                        if s >= self.H * self.M:
                            break
                    #self.started_nodes.append(s)
                    self.started_nodes.append(s)
                    self.ID.append(d)
                    self.earliness.append(e)
                    self.tardiness.append(t)
                print(f'Start: {self.started_nodes} \n End: {self.ID} \n Earliness: {self.earliness} \n Tardiness: {self.tardiness}')
                config.started_nodes = self.started_nodes.copy()
                config.ID = self.ID.copy()
                config.earliness = self.earliness.copy()
                config.tardiness = self.tardiness.copy()

        self.create_tsg_file()
        count = 0
        
        while(count <= num_of_agvs - 1):
            #pdb.set_trace()
            if(isinstance(self.ID, int)):
                self.ID = 3
                self.earliness = 4 if count == 0 else 7
                self.tardiness = 6 if count == 0 else 9
                self.alpha = 1
                self.beta = 1

            self.add_time_windows_constraints()
            assert len(self.ts_edges) == len(self.tsedges), f"Thiếu cạnh ở đâu đó rồi {len(self.ts_edges)} != {len(self.tsedges)}"
            count += 1
        #self.add_restrictions()
        self.gamma = 1
        self.restriction_count = 1
        self.start_ban = 0
        self.end_ban = 2*self.d
        self.restrictions = []
        self.ur = 3
        #pdb.set_trace()
        self.process_restrictions()
        #pdb.set_trace()

    def test_menu(self):
        while True:
            print("======================================")
            print("Nhan (a) de chon file dau vao")
            print("Nhan (b) de in ra ma tran HM")
            print("Nhan (c) de in ra ma tran lien ke Adj")
            print("Nhan (d) de tao ra file TSG.txt")
            print("Nhan (h) de yeu cau nhap ID, earliness, tardiness")
            print("Nhan (j) de cap nhat cac rang buoc ve su xuat hien cua xe")
            print("Nhan cac phim ngoai (a-o) de ket thuc")

            self.use_in_main()
                
    def main_menu(self):
        while True:
            print("======================================")
            print("Nhan (a) de chon file dau vao")
            print("Nhan (b) de in ra ma tran HM")
            print("Nhan (c) de in ra ma tran lien ke Adj")
            print("Nhan (d) de tao ra file TSG.txt")
            print("Nhan (e) de nhap vao ID nguon")
            print("Nhan (f) de kiem tra file")
            print("Nhan (g) de cap nhat file TSG.txt")
            print("Nhan (h) de yeu cau nhap ID, earliness, tardiness")
            print("Nhan (i) de loc ra cac cung cho do thi")
            print("Nhan (j) de cap nhat cac rang buoc ve su xuat hien cua xe")
            print("Nhan (k) de cap nhat cac dong dau cua TSG")
            print("Nhan (l) de loai bo cac dong du thua")
            print("Nhan (m) de loai bo cac dong bi trung lap")
            print("Nhan (o) de giai tim loi giai minimum cho completion time")

            print("Nhan cac phim ngoai (a-o) de ket thuc")

            choice = input("Nhap lua chon cua ban: ").strip().lower()

            if choice == 'a':
                self.getStartedPoints()
                filepath = input("Nhap ten file dau vao: ")
                self.process_input_file(filepath)
            elif choice == 'b':
                self.H = int(input("Nhap vao gia tri H: "))
                self.generate_hm_matrix()
            elif choice == 'c':
                self.d = int(input("Nhap vao gia tri d: "))
                self.generate_adj_matrix()
            elif choice == 'd':
                self.create_tsg_file()
            elif choice == 'e':
                self.query_edges_by_source_id()
            elif choice == 'f':
                self.check_file_conditions()
            elif choice == 'g':
                self.update_file()
            elif choice == 'h':
                self.ID = int(input("Nhập ID của điểm trong không gian: "))
                self.earliness = int(input("Nhập giá trị earliness: "))
                self.tardiness = int(input("Nhập giá trị tardiness: "))
                alpha = input("Nhập alpha (nhấn Enter để lấy giá trị mặc định là 1): ")
                beta = input("Nhập beta (nhấn Enter để lấy giá trị mặc định là 1): ")
                self.alpha = int(alpha) if alpha else 1
                self.beta = int(beta) if beta else 1
                self.add_time_windows_constraints()
            elif choice == 'i':
                self.update_tsg_with_t()
            elif choice == 'j':
                self.add_restrictions()
                self.process_restrictions()
            elif choice == 'k':
                self.add_problem_info()
            elif choice == 'm':
                self.remove_duplicate_lines()
            elif choice == 'l':
                self.remove_redundant_edges()
            elif choice == 'n':
                self.remove_descendant_edges()
            elif choice == 'o':
                self.process_tsg()
            else:
                print("Ket thuc chuong trinh.")
                break

if __name__ == "__main__":
    gp = GraphProcessor()
    #gp.main_menu()
    gp.use_in_main()
